import uuid
import io
import zipfile
from datetime import datetime, timezone, date
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers as xl_numbers
from openpyxl.utils import get_column_letter

from database import get_db
from models.snapshot import ModelSnapshot
from models.deal import Deal
from models.template import Template
from models.extraction import Extraction
from models.classification import ClassificationSession, ClassificationSessionItem
from routers.auth import get_current_user
from models.user import User
from services.excel_exporter import export_populated_model
from services.model_populator import DEFAULT_CELL_MAPPING

router = APIRouter(tags=["models"])


@router.get("/deals/{deal_id}/model")
async def get_model(
    deal_id: uuid.UUID,
    snapshot_id: uuid.UUID = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    deal = await db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if snapshot_id:
        snapshot = await db.get(ModelSnapshot, snapshot_id)
        if not snapshot or snapshot.deal_id != deal_id:
            raise HTTPException(status_code=404, detail="Snapshot not found")
    else:
        result = await db.execute(
            select(ModelSnapshot)
            .where(ModelSnapshot.deal_id == deal_id, ModelSnapshot.is_active == True)
            .order_by(ModelSnapshot.created_at.desc())
        )
        snapshot = result.scalar_one_or_none()

    if not snapshot:
        return {"luckysheet_json": None, "snapshot_id": None}

    return {
        "luckysheet_json": snapshot.luckysheet_json,
        "snapshot_id": str(snapshot.id),
        "snapshot_name": snapshot.snapshot_name,
        "created_at": snapshot.created_at.isoformat(),
    }


@router.put("/deals/{deal_id}/model")
async def save_model(
    deal_id: uuid.UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    deal = await db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    luckysheet_json = body.get("luckysheet_json")
    snapshot_name = body.get("snapshot_name", f"Manual Save {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}")

    if not luckysheet_json:
        raise HTTPException(status_code=400, detail="luckysheet_json is required")

    # Deactivate current active snapshot
    await db.execute(
        update(ModelSnapshot)
        .where(ModelSnapshot.deal_id == deal_id, ModelSnapshot.is_active == True)
        .values(is_active=False)
    )

    snapshot = ModelSnapshot(
        id=uuid.uuid4(),
        deal_id=deal_id,
        snapshot_name=snapshot_name,
        luckysheet_json=luckysheet_json,
        template_id=deal.active_template_id,
        created_at=datetime.now(timezone.utc),
        created_by=current_user.id,
        is_active=True,
    )
    db.add(snapshot)
    deal.last_updated = datetime.now(timezone.utc)
    await db.commit()

    return {"snapshot_id": str(snapshot.id), "snapshot_name": snapshot.snapshot_name}


@router.get("/deals/{deal_id}/model/export")
async def export_model(
    deal_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    deal = await db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    # Template is required for export
    if not deal.active_template_id:
        raise HTTPException(
            status_code=400,
            detail="No template assigned to this deal. Please assign a template before exporting.",
        )
    template = await db.get(Template, deal.active_template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    template_bytes = template.file_data

    # Get latest classification session
    sess_result = await db.execute(
        select(ClassificationSession)
        .where(ClassificationSession.deal_id == deal_id)
        .order_by(ClassificationSession.created_at.desc())
        .limit(1)
    )
    cls_session = sess_result.scalar_one_or_none()
    if not cls_session:
        raise HTTPException(status_code=400, detail="No classification data found. Upload and process a T12 first.")

    # Get session items
    items_result = await db.execute(
        select(ClassificationSessionItem)
        .where(ClassificationSessionItem.session_id == cls_session.id)
        .order_by(ClassificationSessionItem.display_order)
    )
    session_items = items_result.scalars().all()

    # Get rent roll extraction
    rr_result = await db.execute(
        select(Extraction)
        .where(Extraction.deal_id == deal_id, Extraction.document_type == "rent_roll")
        .order_by(Extraction.created_at.desc())
        .limit(1)
    )
    rr_ext = rr_result.scalar_one_or_none()
    rr_data = rr_ext.extracted_data if rr_ext else {}

    # Get OM extraction
    om_result = await db.execute(
        select(Extraction)
        .where(Extraction.deal_id == deal_id, Extraction.document_type == "om")
        .order_by(Extraction.created_at.desc())
        .limit(1)
    )
    om_ext = om_result.scalar_one_or_none()
    om_data = om_ext.extracted_data if om_ext else {}

    # Get T12 period_start from the T12 extraction data
    t12_result = await db.execute(
        select(Extraction)
        .where(Extraction.deal_id == deal_id, Extraction.document_type == "t12")
        .order_by(Extraction.created_at.desc())
        .limit(1)
    )
    t12_ext = t12_result.scalar_one_or_none()
    period_start = t12_ext.extracted_data.get('period_start') if t12_ext else None

    # Write directly from session items to template
    export_bytes = export_populated_model(
        template_bytes=template_bytes,
        session_items=session_items,
        rent_roll_data=rr_data,
        om_data=om_data,
        unit_type_mapping=deal.unit_type_mapping,  # cached from rent roll upload
        period_start=period_start,                 # e.g. "2025-06-01"
    )

    # Detect .xlsm via VBA bin presence
    try:
        with zipfile.ZipFile(io.BytesIO(template_bytes)) as zf:
            is_xlsm = 'xl/vbaProject.bin' in zf.namelist()
    except Exception:
        is_xlsm = False

    ext = '.xlsm' if is_xlsm else '.xlsx'
    mime = ('application/vnd.ms-excel.sheet.macroEnabled.12' if is_xlsm
            else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_name = deal.name.replace(' ', '_')

    return Response(
        content=export_bytes,
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}{ext}"'},
    )


@router.get("/deals/{deal_id}/model/diff")
async def diff_snapshots(
    deal_id: uuid.UUID,
    snapshot_a: uuid.UUID = Query(...),
    snapshot_b: uuid.UUID = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    snap_a = await db.get(ModelSnapshot, snapshot_a)
    snap_b = await db.get(ModelSnapshot, snapshot_b)

    if not snap_a or snap_a.deal_id != deal_id:
        raise HTTPException(status_code=404, detail="Snapshot A not found")
    if not snap_b or snap_b.deal_id != deal_id:
        raise HTTPException(status_code=404, detail="Snapshot B not found")

    # Compare celldata across sheets
    changed_cells = []
    sheets_a = {s["name"]: s.get("celldata", []) for s in snap_a.luckysheet_json.get("sheets", [])}
    sheets_b = {s["name"]: s.get("celldata", []) for s in snap_b.luckysheet_json.get("sheets", [])}

    all_sheet_names = set(sheets_a.keys()) | set(sheets_b.keys())
    for sheet_name in all_sheet_names:
        cells_a = {(c["r"], c["c"]): c.get("v", {}).get("v") for c in sheets_a.get(sheet_name, [])}
        cells_b = {(c["r"], c["c"]): c.get("v", {}).get("v") for c in sheets_b.get(sheet_name, [])}
        all_coords = set(cells_a.keys()) | set(cells_b.keys())
        for coord in all_coords:
            val_a = cells_a.get(coord)
            val_b = cells_b.get(coord)
            if val_a != val_b:
                changed_cells.append({
                    "sheet": sheet_name,
                    "row": coord[0],
                    "col": coord[1],
                    "value_a": val_a,
                    "value_b": val_b,
                })

    return {
        "snapshot_a_id": str(snapshot_a),
        "snapshot_b_id": str(snapshot_b),
        "changed_cells": changed_cells,
        "total_changes": len(changed_cells),
    }


# ── Helpers ────────────────────────────────────────────────────────────────

_BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")
_GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")
_BOLD       = Font(bold=True)
_BOLD_LG    = Font(bold=True, size=14)
_BOLD_CAT   = Font(bold=True, italic=True)
_CAT_FILL   = PatternFill("solid", fgColor="E2EFDA")

_NUM_FMT    = '#,##0.00'
_CCY_FMT    = '$#,##0.00'

_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_CATEGORY_LABELS = {
    "MarketRent":  "MARKET RENT",
    "LTL":         "LOSS TO LEASE",
    "Vacancy":     "VACANCY",
    "Concessions": "CONCESSIONS",
    "BadDebt":     "BAD DEBT",
    "RUBSInc":     "RUBS INCOME",
    "RetailInc":   "RETAIL INCOME",
    "OtherInc":    "OTHER INCOME",
    "Payroll":     "PAYROLL AND BENEFITS",
    "MgmtFee":     "MANAGEMENT FEES",
    "Landscaping": "LANDSCAPING",
    "Repairs":     "REPAIRS AND MAINTENANCE",
    "Turnover":    "TURNOVER",
    "Utilities":   "UTILITIES",
    "SecurityLife": "SECURITY / LIFE SAFETY",
    "Advert":      "ADVERTISING",
    "Admin":       "ADMINISTRATIVE",
    "Insurance":   "INSURANCE",
    "PropTax":     "PROPERTY TAXES",
    "MiscExp":     "MISCELLANEOUS EXPENSE",
    "CapEx":       "CAPITAL EXPENDITURES",
}


def _add_months(d: date, n: int) -> date:
    month = d.month - 1 + n
    year = d.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)


def _period_month_sequence(period_start: str | None) -> list[tuple[str, str]]:
    """
    Returns list of (month_abbr, col_label) for 12 months starting at period_start.
    e.g. period_start="2025-06-01" → [("Jun","Jun-25"), ("Jul","Jul-25"), ..., ("May","May-26")]
    Falls back to Jan–Dec if period_start is missing/unparseable.
    """
    try:
        start = date.fromisoformat(period_start) if period_start else date(2025, 1, 1)
    except Exception:
        start = date(2025, 1, 1)

    pairs = []
    for i in range(12):
        d = _add_months(start, i)
        abbr = _MONTH_ABBR[d.month - 1]
        label = f"{abbr}-{str(d.year)[2:]}"
        pairs.append((abbr, label))
    return pairs


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── T12 Export ─────────────────────────────────────────────────────────────

@router.get("/deals/{deal_id}/export/t12")
async def export_t12(
    deal_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    deal = await db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    t12_result = await db.execute(
        select(Extraction)
        .where(Extraction.deal_id == deal_id, Extraction.document_type == "t12")
        .order_by(Extraction.created_at.desc())
        .limit(1)
    )
    t12_ext = t12_result.scalar_one_or_none()
    period_start = t12_ext.extracted_data.get("period_start") if t12_ext else None

    sess_result = await db.execute(
        select(ClassificationSession)
        .where(ClassificationSession.deal_id == deal_id)
        .order_by(ClassificationSession.created_at.desc())
        .limit(1)
    )
    cls_session = sess_result.scalar_one_or_none()
    if not cls_session:
        raise HTTPException(status_code=400, detail="No classification data found. Upload and process a T12 first.")

    items_result = await db.execute(
        select(ClassificationSessionItem)
        .where(ClassificationSessionItem.session_id == cls_session.id)
        .order_by(ClassificationSessionItem.display_order)
    )
    session_items = items_result.scalars().all()

    month_pairs = _period_month_sequence(period_start)  # [(abbr, label), ...]

    wb = Workbook()
    ws = wb.active
    ws.title = "T12 Income Statement"

    # Row 1 – deal name
    ws.cell(1, 1, deal.name).font = _BOLD_LG
    # Row 2 – title
    ws.cell(2, 1, "Income Statement (12 months)").font = _BOLD
    # Row 3 – period range
    if month_pairs:
        period_str = f"Period: {month_pairs[0][1]} – {month_pairs[-1][1]}"
        ws.cell(3, 1, period_str)
    # Row 4 – blank
    # Row 5 – header
    ws.cell(5, 1, "Category").font = _BOLD
    ws.cell(5, 1).fill = _BLUE_FILL
    for col_idx, (_, label) in enumerate(month_pairs, start=2):
        c = ws.cell(5, col_idx, label)
        c.font = _BOLD
        c.fill = _BLUE_FILL
        c.alignment = Alignment(horizontal="right")
    ws.cell(5, 14, "Total").font = _BOLD
    ws.cell(5, 14).fill = _BLUE_FILL
    ws.cell(5, 14).alignment = Alignment(horizontal="right")

    # Group items by final_category
    from collections import OrderedDict
    groups: OrderedDict[str, list] = OrderedDict()
    for item in session_items:
        cat = item.final_category or "Uncategorized"
        groups.setdefault(cat, []).append(item)

    data_row = 6
    shade = False
    for cat, cat_items in groups.items():
        # Category header row
        label = _CATEGORY_LABELS.get(cat, cat.upper())
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=14)
        c = ws.cell(data_row, 1, label)
        c.font = _BOLD_CAT
        c.fill = _CAT_FILL
        data_row += 1

        for item in cat_items:
            mv = item.monthly_values or {}
            row_fill = _GRAY_FILL if shade else None
            ws.cell(data_row, 1, item.line_item_name)
            if row_fill:
                ws.cell(data_row, 1).fill = row_fill

            for col_idx, (abbr, _) in enumerate(month_pairs, start=2):
                val = mv.get(abbr) or 0
                c = ws.cell(data_row, col_idx, val)
                c.number_format = _NUM_FMT
                c.alignment = Alignment(horizontal="right")
                if row_fill:
                    c.fill = row_fill

            total_cell = ws.cell(data_row, 14, item.trailing_total or 0)
            total_cell.number_format = _NUM_FMT
            total_cell.alignment = Alignment(horizontal="right")
            if row_fill:
                total_cell.fill = row_fill

            shade = not shade
            data_row += 1

        # Blank row between groups
        data_row += 1

    _set_col_widths(ws, [36] + [11] * 12 + [13])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = deal.name.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_T12.xlsx"'},
    )


# ── Rent Roll Export ────────────────────────────────────────────────────────

@router.get("/deals/{deal_id}/export/rentroll")
async def export_rentroll(
    deal_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    deal = await db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    rr_result = await db.execute(
        select(Extraction)
        .where(Extraction.deal_id == deal_id, Extraction.document_type == "rent_roll")
        .order_by(Extraction.created_at.desc())
        .limit(1)
    )
    rr_ext = rr_result.scalar_one_or_none()
    if not rr_ext:
        raise HTTPException(status_code=400, detail="No rent roll data found. Upload a rent roll document first.")

    rr_data = rr_ext.extracted_data or {}
    as_of_date = rr_data.get("as_of_date") or ""
    unit_mix = rr_data.get("unit_mix") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Rent Roll"

    # Row 1 – deal name
    ws.cell(1, 1, deal.name).font = _BOLD_LG
    # Row 2 – title
    ws.cell(2, 1, "Rent Roll with Lease Charges").font = _BOLD
    # Row 3 – as-of date
    ws.cell(3, 1, f"As Of: {as_of_date}")
    # Row 4 – blank

    # Row 5 – headers
    HEADERS = ["Unit", "Unit Type", "SF", "Status", "Market Rent",
               "Charge Code", "Amount", "Move In", "Lease Expiration"]
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(5, col_idx, h)
        c.font = _BOLD
        c.fill = _BLUE_FILL

    data_row = 6
    shade = False
    for unit in unit_mix:
        unit_number  = unit.get("unit_number") or ""
        unit_type    = unit.get("unit_type") or ""
        sf           = unit.get("sf")
        status       = unit.get("status") or ""
        market_rent  = unit.get("market_rent")
        move_in      = unit.get("move_in") or ""
        lease_exp    = unit.get("lease_expiration") or ""
        charges      = unit.get("charges") or {}

        row_fill = _GRAY_FILL if shade else None

        if status == "vacant" or not charges:
            # One row, no charge detail
            row_vals = [unit_number, unit_type, sf, status.upper(),
                        market_rent, "", "", move_in, lease_exp]
            for col_idx, val in enumerate(row_vals, start=1):
                c = ws.cell(data_row, col_idx, val)
                if col_idx == 5 and val is not None:
                    c.number_format = _CCY_FMT
                if row_fill:
                    c.fill = row_fill
            data_row += 1
            shade = not shade
        else:
            first = True
            for charge_code, amount in charges.items():
                if first:
                    row_vals = [unit_number, unit_type, sf, status.upper(),
                                market_rent, charge_code, amount, move_in, lease_exp]
                else:
                    # continuation rows: blank out unit-level fields
                    row_vals = ["", "", "", "", "", charge_code, amount, "", ""]

                for col_idx, val in enumerate(row_vals, start=1):
                    c = ws.cell(data_row, col_idx, val)
                    if col_idx == 5 and val not in ("", None):
                        c.number_format = _CCY_FMT
                    if col_idx == 7 and val not in ("", None):
                        c.number_format = _CCY_FMT
                    if row_fill:
                        c.fill = row_fill
                data_row += 1
                first = False
            shade = not shade

    _set_col_widths(ws, [8, 12, 7, 10, 14, 14, 12, 12, 18])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = deal.name.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_RentRoll.xlsx"'},
    )
