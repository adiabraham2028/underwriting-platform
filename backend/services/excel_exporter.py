import io
import re
import json
import zipfile
import logging
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Unit type mapping utilities (Part 1)
# ---------------------------------------------------------------------------

def get_template_unit_type_map(template_bytes: bytes) -> dict:
    """
    Read the template Rent Roll tab rows 7-30 and return
    {template_code: description}, e.g. {"1B D": "1 BR / 1 BA"}.
    """
    try:
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True, read_only=True)
    except Exception as e:
        logger.warning(f"get_template_unit_type_map: could not open template: {e}")
        return {}
    ws = None
    for name in wb.sheetnames:
        if re.sub(r'[\s\-_]', '', name).upper() == 'RENTROLL':
            ws = wb[name]
            break
    if not ws:
        return {}
    unit_types = {}
    for row in ws.iter_rows(min_row=7, max_row=30, min_col=1, max_col=2, values_only=True):
        code, desc = row
        if code and desc and str(code) not in ('UnitType', 'Unit'):
            unit_types[str(code)] = str(desc)
    return unit_types


def parse_conam_beds_baths(unit_type: str) -> tuple:
    """G12_22AR → (2, 2).  G12_11AR → (1, 1).  G12_32AR → (3, 2)."""
    m = re.search(r'_(\d)(\d)[A-Z]', str(unit_type))
    if m:
        return int(m.group(1)), int(m.group(2))
    return (0, 0)


def map_conam_to_template_codes(unit_mix: list, template_unit_types: dict) -> dict:
    """
    Local mapping for CONAM format — no Claude needed.
    Returns {conam_code: template_code}.
    """
    template_by_beds: dict = defaultdict(list)
    for code, desc in template_unit_types.items():
        m = re.search(r'(\d+)\s*BR\s*/\s*(\d+)\s*BA', str(desc), re.IGNORECASE)
        if m:
            key = (int(m.group(1)), int(m.group(2)))
            template_by_beds[key].append(code)

    mapping: dict = {}
    seen: set = set()
    for unit in unit_mix:
        ut = str(unit.get('unit_type', ''))
        if ut in seen:
            continue
        seen.add(ut)
        beds, baths = parse_conam_beds_baths(ut)
        candidates = template_by_beds.get((beds, baths), [])
        mapping[ut] = candidates[0] if candidates else ut
    return mapping


async def map_unit_types_with_claude(
    rent_roll_types: list,
    template_unit_types: dict,
    llm_service,
) -> dict:
    """
    Claude-based mapping for non-CONAM formats.
    One API call per deal; result is cached on the deal.
    """
    system = (
        "You are a real estate analyst. Map rent roll unit type codes "
        "to template codes by bed/bath/sf. Return ONLY valid JSON, no markdown."
    )
    content = f"""
Rent roll unit types:
{json.dumps(rent_roll_types, indent=2)}

Available template codes:
{json.dumps(template_unit_types, indent=2)}

Map each rent roll code to the single best matching template code.
If no close match exists, use the first template code with the same bed count.

Return ONLY:
{{"mapping": {{"RENT_ROLL_CODE": "TEMPLATE_CODE"}}}}
"""
    result = await llm_service.complete_json_with_retry(system, content, max_tokens=512)
    return result.get("mapping", {})


# ---------------------------------------------------------------------------

MONTH_ORDER = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _sort_months(month_keys: list[str]) -> list[str]:
    """Sort month abbreviation keys into calendar order."""
    return sorted(month_keys, key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)


def _parse_beds_baths(unit_type: str) -> tuple[float, float]:
    """
    Parse bed/bath from CONAM unit type code.
    Format: G[bldg][beds][baths][subtype], e.g. G12_11AR → 1 bed 1 bath
    """
    if not unit_type:
        return (0, 0)
    match = re.search(r'_(\d)(\d)[A-Z]', str(unit_type))
    if match:
        return (float(match.group(1)), float(match.group(2)))
    return (0, 0)


def _write_t12(ws, session_items: list, period_start: str = None):
    """
    Find the CATEGORY header row in the T12 sheet and write one row per
    session item immediately below it.

    period_start: ISO date string like "2025-06-01". When provided, writes to
    the template's B16 cell (the period start input that drives all EOMONTH
    month-header formulas). Also updates the A15 period label.
    """
    # Write period start date to B16 before touching anything else.
    # B16 is the key input cell — EOMONTH formulas in C16..M16 cascade from it.
    if period_start:
        try:
            from datetime import datetime as _dt
            ps = _dt.fromisoformat(period_start)
            b16 = ws.cell(row=16, column=2)
            if not str(b16.value or '').startswith('='):
                b16.value = ps
            # Also update the period label in A15 if it's plain text
            a15 = ws.cell(row=15, column=1)
            if a15.value and not str(a15.value).startswith('='):
                end_month = (ps.month + 10) % 12 + 1
                end_year  = ps.year + 1 if ps.month > 2 else ps.year
                a15.value = (f"Period = {ps.strftime('%b %y')} - "
                             f"{_dt(end_year, end_month, 1).strftime('%b %y')}")
        except Exception as e:
            logger.warning(f"Could not write T12 period start: {e}")

    category_row = None
    category_col = None
    for row in ws.iter_rows(max_row=50):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == 'CATEGORY':
                category_row = cell.row
                category_col = cell.column
                break
        if category_row:
            break

    if not category_row:
        category_row = 7
        category_col = 1

    # Build month → column mapping from the CATEGORY header row
    month_col_map: dict[str, int] = {}
    for cell in ws[category_row]:
        if cell.column <= category_col:
            continue
        if cell.value:
            val = str(cell.value).strip()
            abbr = val[:3].capitalize()
            if abbr in MONTH_ORDER:
                month_col_map[abbr] = cell.column

    data_start = category_row + 1

    # Clear stale data below the header (preserve header and above)
    for row_idx in range(data_start, data_start + 400):
        for col_idx in range(category_col, category_col + 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and not str(cell.value).startswith('='):
                cell.value = None

    sorted_items = sorted(
        [i for i in session_items if i.final_category and i.final_category != 'SKIP'],
        key=lambda x: x.display_order or 0,
    )

    for i, item in enumerate(sorted_items):
        row_idx = data_start + i
        monthly = item.monthly_values or {}

        # Column A (category_col): line item name
        ws.cell(row=row_idx, column=category_col).value = item.line_item_name

        if month_col_map:
            # Write to matched month columns from template header
            for abbr, col in month_col_map.items():
                ws.cell(row=row_idx, column=col).value = monthly.get(abbr, 0)
        else:
            # Fallback: write in calendar order starting at category_col + 1
            sorted_months = _sort_months(list(monthly.keys()))
            for j, month in enumerate(sorted_months[:12]):
                ws.cell(row=row_idx, column=category_col + 1 + j).value = monthly.get(month, 0)

        # Trailing total (14th column after category_col)
        ws.cell(row=row_idx, column=category_col + 13).value = item.trailing_total or 0

        # Assignment — the category name the SUMIF formulas aggregate on
        ws.cell(row=row_idx, column=category_col + 14).value = item.final_category


def _write_rent_roll(ws, rr_data: dict, unit_type_mapping: dict):
    """
    Write rent roll into the RR Excel table.
    One row per CHARGE (RENT and GARAGE on separate rows).
    Clears rows 31+ before writing; maps CONAM codes via unit_type_mapping.
    """
    unit_mix = rr_data.get('unit_mix', [])
    if not unit_mix:
        return 0

    # Write "As Of" date to A3 (row 3, col 1).
    # The template has a stale hardcoded date here; overwrite with the actual
    # report date extracted from the source document.
    as_of_raw = rr_data.get('as_of_date')
    if as_of_raw:
        try:
            from datetime import datetime as _dt
            as_of = _dt.fromisoformat(as_of_raw) if isinstance(as_of_raw, str) else as_of_raw
            a3 = ws.cell(row=3, column=1)
            if not str(a3.value or '').startswith('='):
                a3.value = as_of
        except Exception as e:
            logger.warning(f"Could not write Rent Roll As Of date: {e}")

    DATA_START = 31

    # Clear existing data rows (A–L), skip formula cells
    for row_idx in range(DATA_START, DATA_START + 700):
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).startswith('='):
                continue
            cell.value = None

    current_row = DATA_START

    for unit in unit_mix:
        raw_type    = str(unit.get('unit_type', ''))
        tmpl_type   = unit_type_mapping.get(raw_type, raw_type)
        sf          = unit.get('sf')
        market_rent = unit.get('market_rent')
        move_in     = unit.get('move_in')
        lease_exp   = unit.get('lease_expiration')
        status      = unit.get('status', 'occupied').capitalize()
        resident_id = unit.get('resident_id', '')

        charges = unit.get('charges') or {}
        if not charges:
            charges = {'RENT': unit.get('base_rent', 0)}

        for charge_code, amount in charges.items():
            ws.cell(row=current_row, column=1).value  = unit.get('unit_number')
            ws.cell(row=current_row, column=2).value  = tmpl_type
            ws.cell(row=current_row, column=3).value  = sf
            ws.cell(row=current_row, column=4).value  = status
            ws.cell(row=current_row, column=5).value  = resident_id
            ws.cell(row=current_row, column=6).value  = market_rent
            ws.cell(row=current_row, column=7).value  = charge_code
            ws.cell(row=current_row, column=8).value  = amount
            ws.cell(row=current_row, column=11).value = move_in
            ws.cell(row=current_row, column=12).value = lease_exp
            current_row += 1

    return current_row - DATA_START


def _write_summary_fields(ws, rr_data: dict, om_data: dict):
    """Write key property fields to the summary tab, skipping formula cells."""
    def safe_write(addr, value):
        if value is None:
            return
        cell = ws[addr]
        if cell.value and str(cell.value).startswith('='):
            return
        cell.value = value

    # Try common locations — property name
    for addr in ['B2', 'C2', 'D2']:
        cell = ws[addr]
        if not cell.value or not str(cell.value).startswith('='):
            safe_write(addr, om_data.get('property_name'))
            break

    # Units
    for addr in ['B6', 'C6', 'B5']:
        cell = ws[addr]
        if not cell.value or not str(cell.value).startswith('='):
            safe_write(addr, rr_data.get('total_units'))
            break


def export_populated_model(
    template_bytes: bytes,
    session_items: list,           # ClassificationSessionItem ORM objects
    rent_roll_data: dict,          # extracted_data from rent roll extraction
    om_data: dict,                 # extracted_data from OM (may be empty)
    unit_type_mapping: dict = None,  # cached from deal.unit_type_mapping
    period_start: str = None,      # ISO date "YYYY-MM-DD" for T12 B16 cell
) -> bytes:
    """
    Load the .xlsm template, write T12 line items and Rent Roll units
    directly using openpyxl. Return modified .xlsm bytes.
    """
    try:
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    except Exception as e:
        logger.error(f"Failed to load template: {e}")
        raise

    sheet_names = wb.sheetnames

    # Find T12 sheet
    t12_sheet = None
    for name in sheet_names:
        if re.sub(r'[\s\-_]', '', name).upper() == 'T12':
            t12_sheet = wb[name]
            break

    # Find Rent Roll sheet
    rr_sheet = None
    for name in sheet_names:
        if re.sub(r'[\s\-_]', '', name).upper() == 'RENTROLL':
            rr_sheet = wb[name]
            break

    if t12_sheet and session_items:
        _write_t12(t12_sheet, session_items, period_start=period_start)

    if rr_sheet and rent_roll_data.get('unit_mix'):
        # Resolve unit type mapping: use cached, or compute locally
        if not unit_type_mapping:
            tmpl_types = get_template_unit_type_map(template_bytes)
            unit_type_mapping = map_conam_to_template_codes(
                rent_roll_data['unit_mix'], tmpl_types
            )
        rows_written = _write_rent_roll(rr_sheet, rent_roll_data, unit_type_mapping)
        logger.info(f"Rent roll: {rows_written} charge rows written to template")

    # Write summary fields
    for summary_name in ['Exec Summary', 'TCP Summary', 'Summary', 'Executive Summary']:
        if summary_name in sheet_names:
            _write_summary_fields(wb[summary_name], rent_roll_data, om_data)
            break

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Legacy helpers (kept for backward compat and the default template)
# ---------------------------------------------------------------------------

def export_to_excel(template_bytes: bytes, luckysheet_json: dict, cell_mapping: dict) -> bytes:
    """Legacy: write Luckysheet celldata back into the template via openpyxl."""
    try:
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    except Exception as e:
        logger.warning(f"Could not load template ({e}), creating new workbook")
        wb = openpyxl.Workbook()

    for sheet_data in luckysheet_json.get("sheets", []):
        sheet_name = sheet_data.get("name", "Sheet")
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        for entry in sheet_data.get("celldata", []):
            row = entry.get("r", 0) + 1
            col = entry.get("c", 0) + 1
            value = (entry.get("v") or {}).get("v")
            existing = ws.cell(row=row, column=col)
            if existing.value and str(existing.value).startswith("="):
                continue
            ws.cell(row=row, column=col, value=value)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def create_default_template() -> bytes:
    """Create the simple 7-tab default template (used when no client template is uploaded)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_s = wb.create_sheet("Summary")
    for cell, val in [("A1","PROPERTY SUMMARY"),("A2","Property Name"),("A3","Address"),
                      ("A4","City/State/Zip"),("A6","Total Units"),("A7","Total SF"),
                      ("A8","Year Built"),("A10","Asking Price"),("A11","Current Occupancy"),
                      ("A12","Current NOI"),("A13","Cap Rate (In-Place)")]:
        ws_s[cell] = val

    ws_rr = wb.create_sheet("Rent Roll")
    ws_rr["A1"] = "RENT ROLL SUMMARY"
    ws_rr["A3"] = "Occupancy Rate"
    ws_rr["A4"] = "Avg Current Rent"
    ws_rr["A5"] = "Avg Market Rent"
    for ci, h in enumerate(["Unit #","Beds","Baths","SF","Current Rent",
                              "Market Rent","Lease Start","Lease End","Status"], 1):
        ws_rr.cell(row=7, column=ci, value=h)

    ws_t = wb.create_sheet("T-12")
    ws_t["A1"] = "TRAILING 12-MONTH OPERATING STATEMENT"
    for ci, h in enumerate(["CATEGORY","Jan","Feb","Mar","Apr","May","Jun",
                              "Jul","Aug","Sep","Oct","Nov","Dec","Total","Assignment"], 1):
        ws_t.cell(row=7, column=ci, value=h)

    for sheet_name in ["Pro Forma", "Debt", "Returns", "Assumptions"]:
        wb.create_sheet(sheet_name)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
