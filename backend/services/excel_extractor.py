import openpyxl
import io
import logging

logger = logging.getLogger(__name__)


def extract_excel_text(file_bytes: bytes) -> str:
    """Extract text from an Excel file."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            texts.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    texts.append("\t".join(row_values))
        return "\n".join(texts)
    except Exception as e:
        logger.error(f"Error extracting Excel text: {e}")
        raise


def get_excel_structure(file_bytes: bytes) -> dict:
    """Get the structure of an Excel file for template mapping."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        structure = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells = {}
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value is not None:
                        cells[cell.coordinate] = {
                            "value": str(cell.value),
                            "row": cell.row,
                            "column": cell.column_letter,
                        }
            structure[sheet_name] = {"cells": cells, "max_row": ws.max_row, "max_col": ws.max_column}
        return structure
    except Exception as e:
        logger.error(f"Error getting Excel structure: {e}")
        raise


def extract_excel_rows(file_bytes: bytes, sheet_index: int = 0) -> list[dict]:
    """Extract rows as list of dicts from the first (or specified) sheet."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if not any(v is not None for v in row):
                continue
            result.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return result
    except Exception as e:
        logger.error(f"Error extracting Excel rows: {e}")
        return []
